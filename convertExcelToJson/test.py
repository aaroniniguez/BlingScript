import openpyxl
import collections
Camera = "RED Cameras"
wb = openpyxl.load_workbook("data.xlsx")
sheet = wb.get_sheet_by_name(Camera)
compression = []
myrange = range(612,802)+range(848,1563)+range(1609,2031)
for i in myrange:
	if sheet.cell(row=i,column=1).value is not None:
		compressionCur = sheet.cell(row=i,column=3).value
		if compressionCur == "ProRes 422":
			compressionCur = "ProRes422"
		resolution = sheet.cell(row=i,column=2).value.split("\n")[0]
		easyName = sheet.cell(row=i,column=2).value.split("\n")[1].replace("(","").replace(")","")
		rateId = str(sheet.cell(row=i,column=4).value).replace("p","").replace(".0","")
		rateName = str(sheet.cell(row=i,column=6).value*8).replace(".0","")
		idExist = False
		for item in compression:
			if item["id"] == compressionCur:
				resolExist = False
				idExist = True
				for resol in item["res"]:
					if resol["id"] == resolution:
						resol["rate"].append({"id":rateId,"name":rateName})
						resolExist = True
				if not resolExist:
					mydict = collections.OrderedDict()
					mydict["id"] = resolution
					mydict["easyName"] = easyName
					mydict["rate"] = [{"id":rateId,"name":rateName}]
					#item["res"].append({"id":resolution,"easyName":easyName,"rate":[{"id":rateId,"name":rateName}]})
					item["res"].append(mydict)
		if not idExist:
			mydict = collections.OrderedDict()
			mydict["id"] = resolution
			mydict["easyName"] = easyName
			mydict["rate"] = [{"id":rateId,"name":rateName}]
			mydict2 = collections.OrderedDict()
			mydict2["id"] = compressionCur
			mydict2["res"] = [mydict]
			#compression.append({"id":compressionCur,"res":[{"id":resolution,"easyName":easyName,"rate":[{"id":rateId,"name":rateName}]}]})
			#compression.append({"id":compressionCur,"res":[mydict]})
			compression.append(mydict2)
for i in compression:
	print i["id"]
import json
with open("text", "w") as outfile:
    json.dump(compression, outfile, indent=4)
