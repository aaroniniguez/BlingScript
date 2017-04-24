# Author: Aaron Iniguez
# convert excel values into json file
import openpyxl
import collections
sheetName = "Sony Cameras"
Camera = "Sony AS7"
wb = openpyxl.load_workbook("data.xlsx")
sheet = wb.get_sheet_by_name(sheetName)
codecs = []
myrange = range(54,64)
for i in myrange:
	if sheet.cell(row=i,column=1).value is not None:
		compression = sheet.cell(row=i,column=3).value.replace("\n","")
		if compression == "ProRes 422":
			compression = "ProRes422"
		resInfo = sheet.cell(row=i,column=2).value.split("\n")
		resolution = sheet.cell(row=i,column=2).value.split("\n")[0]
		if len(resInfo) > 1:
			easyName = sheet.cell(row=i,column=2).value.split("\n")[1].replace("(","").replace(")","")
		else:
			easyName = ""
		rateId = str(sheet.cell(row=i,column=4).value).replace("p","").replace(".0","")
		rateName = str(sheet.cell(row=i,column=5).value).replace(".0","")
		idExist = False
		for item in codecs:
			if item["id"] == compression:
				resolExist = False
				idExist = True
				for resol in item["res"]:
					if resol["id"] == resolution:
						resol["rate"].append({"id":rateId,"name":rateName})
						resolExist = True
				if not resolExist:
					mydict = collections.OrderedDict()
					mydict["id"] = resolution
					mydict["easyName"] = easyName
					mydict["rate"] = [{"id":rateId,"name":rateName}]
					item["res"].append(mydict)
		if not idExist:
			mydict = collections.OrderedDict()
			mydict["id"] = resolution
			mydict["easyName"] = easyName
			mydict["rate"] = [{"id":rateId,"name":rateName}]
			mydict2 = collections.OrderedDict()
			mydict2["id"] = compression
			mydict2["res"] = [mydict]
			codecs.append(mydict2)
for i in codecs:
	print i["id"]
import json
outPut = {"id":Camera,
		"codecs":codecs}
with open("text", "w") as outfile:
    json.dump(outPut, outfile, indent=4)
